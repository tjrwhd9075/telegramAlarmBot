
# #하위폴더 files에 저장된 모든 파일들의 이름을 리스트로 저장하고 출력
import os
path = "C:/Users/seokjong_2/Desktop/OneDrive/Documents/"
# file_list = os.listdir(path)
# print(file_list)

# #파일열고, 시트에서 값 가져오기
from openpyxl import load_workbook
# results = []
# for file_name_raw in file_list:

file_name = path + "포트폴리오2.xlsm"
wb = load_workbook(filename=file_name,read_only=False, data_only=True)
ws = wb["한국"]
print(ws['c3'].value)
wb.save('test.xlsx')

# #셀 좌표로 값 출력
# print(ws.cell(1,2).value)

# ws = wb.active   #현재 활성화된 워크시트
#     result = []
#     result.append(file_name_raw)
#     result.append(ws['C2'].value)
#     result.append(ws['C3'].value)
#     result.append(ws['E2'].value)
#     result.append(ws['E3'].value)
#     results.append(result)
# print(results)

# # 새로운 파일 생성, 저장
# from openpyxl import Workbook
# wb = Workbook()
# ws = wb.active
# for i in results:
#     ws.append(i)
# wb.save("results.xlsx")


